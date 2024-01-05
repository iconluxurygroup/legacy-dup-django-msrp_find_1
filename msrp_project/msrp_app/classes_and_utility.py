# Import statements
import requests
from bs4 import BeautifulSoup
import json
import csv
import re
import logging
import random
from urllib.parse import urlparse
import time
from openpyxl import load_workbook
import datetime
class Product:
    def __init__(self, input_sku, brand):
        self.input_sku = input_sku
        self.brand = brand
        self.variations = []
        self.title = None
        self.images = []
        self.prices = []
        self.currency = None
        self.url = None
        self.description = None
        self.source_type = None  # New attribute to store the source type
        self.seller = None  # New attribute to store the seller information
        self.excel_row_number=None
    def add_variation(self, variation):
        self.variations.append(variation)

    def set_details(self, title, images, prices, currency, url, description, seller):
        self.title = title
        self.images = images
        self.prices = prices
        self.currency = currency
        self.url = url
        self.description = description
        self.seller = seller  # Set the seller information

    def is_complete(self):
        return bool(self.prices and self.url)


class BrandSettings:
    def __init__(self, settings):
        self.settings = settings

    def get_rules_for_brand(self, brand_name):
        for rule in self.settings['brand_rules']:
            if brand_name in rule['names']:
                return rule
        return None
    

class SKUManager:
    def __init__(self, brand_settings):
        self.brand_settings = brand_settings
    def generate_variations(self, input_sku, brand_rule):
        brand_variations = self.handle_brand_sku(input_sku, brand_rule)
        blind_variations = self.handle_sku(input_sku, brand_rule)
        return brand_variations + blind_variations# + [input_sku]

    def handle_brand_sku(self, sku, brand_rule):
        cleaned_sku = self.clean_sku(sku)
        sku_format = brand_rule['sku_format']
        base_parts = sku_format['base']
        base_separator = sku_format.get('base_separator', '')
        color_separator = sku_format.get('color_separator', '')

        article_length = int(base_parts['article'][0].split(',')[0])
        model_length = int(base_parts['model'][0].split(',')[0])
        color_length = int(sku_format['color_extension'][0].split(',')[0])

        article_part = cleaned_sku[:article_length]
        model_part = cleaned_sku[article_length:article_length + model_length]
        color_part = cleaned_sku[article_length + model_length:article_length + model_length + color_length]

        # Order: Brand Format with Color, Brand Format without Color
        return [
        article_part + base_separator + model_part + color_separator + color_part,
        article_part + base_separator + model_part
    ]

    def handle_sku(self, sku, brand_rule):
        cleaned_sku = self.clean_sku(sku)
        sku_format = brand_rule['sku_format']

        article_length = int(sku_format['base']['article'][0].split(',')[0])
        model_length = int(sku_format['base']['model'][0].split(',')[0])
        color_length = int(sku_format['color_extension'][0].split(',')[0])

        article_part = cleaned_sku[:article_length]
        model_part = cleaned_sku[article_length:article_length + model_length]
        color_part = cleaned_sku[article_length + model_length:article_length + model_length + color_length]

        # Order: No space (Article Model Color), Space (Article Model Color), No space (Article Model), Space (Article Model)
        return [
            article_part + model_part + color_part, 
            article_part + ' ' + model_part + ' ' + color_part,
            article_part + model_part,
            article_part + ' ' + model_part
        ]


    @staticmethod
    def clean_sku(sku):
        sku = str(sku)
        logging.info(f"Cleaning SKU: {sku}")
        cleaned = re.sub(r'[^a-zA-Z0-9]', '', sku)
        logging.info(f"Cleaned SKU: {cleaned}")
        return cleaned


class DataFetcher:
    @staticmethod
    def parse_google_results(html_content):
        soup = BeautifulSoup(html_content, 'html.parser')
        results = []
        for g in soup.find_all('div', class_='g'):
            links = g.find_all('a')
            if links and 'href' in links[0].attrs:  # check if 'href' attribute exists
                results.append(links[0]['href'])
        return results
    
    def extract_product_schema(self, html_content):
        product_schemas = []  # List to store all found product schemas

        try:
            soup = BeautifulSoup(html_content, 'html.parser')
            schema_tags = soup.find_all('script', {'type': 'application/ld+json'})

            for tag in schema_tags:
                try:
                    data = json.loads(tag.text)
                    if data.get('@type') == 'Product':
                        # Log the raw product schema for debugging
                        logging.debug("Raw Product Schema: %s", json.dumps(data, indent=4))
                        product_schemas.append(data)
                except json.JSONDecodeError:
                    continue

            if not product_schemas:
                logging.warning("No Product schema found in the HTML content.")
                return None

            return product_schemas
        except Exception as e:
            logging.error(f"Error extracting product schemas from HTML: {e}")
            return None
    def parse_product_schemas(self, product_schemas):
        parsed_products = []

        for schema in product_schemas:
            if schema.get('@type') == 'Product':
                # Attempt to get the URL from the top level; if not found, check in 'offers'
                url = schema.get('url', '') or schema.get('offers', {}).get('url', '')
#!
                seller = schema.get('offers', {}).get('seller', '')
                if isinstance(seller, dict) and 'name' in seller:
                    seller = seller['name']
                elif isinstance(seller, list) and len(seller) > 0 and 'name' in seller[0]:
                    seller = seller[0]['name']
#!
                product_details = {
                    'title': schema.get('name', ''),
                    'images': [schema.get('image')] if schema.get('image') else [],
                    'prices': [float(schema['offers']['price'].replace('$', ''))] if schema.get('offers', {}).get('price') else [],
#!
                    # 'prices': [float(schema['offers']['price'])] if schema.get('offers', {}).get('price') else [],
                    'currency': schema.get('offers', {}).get('priceCurrency', ''),
                    'url': url,
                    'description': schema.get('description', ''),
                    'seller': seller.lower(),  # Include the seller information as a string
                    # Add more fields as needed
                }
                parsed_products.append(product_details)
#!
        return parsed_products
    #def parse_product_schemas(self, product_schemas):
    #    parsed_products = []
#
    #    for schema in product_schemas:
    #        print(schema)
    #        if schema.get('@type') == 'Product':
    #            url = schema.get('url', '') or schema.get('offers', {}).get('url', '')
#
    #            # Improved seller extraction
    #            seller = ''
    #            offers = schema.get('offers', {})
    #            #!FIX
    #            offers = offers.get('offers', {})
    #            print(f"OFFERS: {type(offers)}")
    #            print(offers)
    #            if isinstance(offers, dict):
    #                print(f"OFFERS 1:")
    #                seller_info = offers.get('seller', {})
    #                print(f"SELLER: {seller_info}")
    #                if isinstance(seller_info, dict):
    #                    print(f"OFFERS 2:")
    #                    seller = seller_info.get('name', '')
    #                    print(f"SELLER: {seller}")
    #                elif isinstance(seller_info, list) and seller_info:
    #                    print(f"OFFERS 3:")
    #                    seller = seller_info[0].get('name', '') if isinstance(seller_info[0], dict) else ''
    #                    print(f"SELLER: {seller}")
    #            elif isinstance(offers, list) and offers:
    #                offers = offers[0]
    #                print(f"OFFERS: {type(offers)}")
    #                print(offers)
    #                if isinstance(offers, dict):
    #                    print(f"OFFERS 1:")
    #                    seller_info = offers.get('seller', {})
    #                    print(f"SELLER: {seller_info}")
    #                    if isinstance(seller_info, dict):
    #                        print(f"OFFERS 2:")
    #                        seller = seller_info.get('name', '')
    #                        print(f"SELLER: {seller}")
    #                    elif isinstance(seller_info, list) and seller_info:
    #                        print(f"OFFERS 3:")
    #                        seller = seller_info[0].get('name', '') if isinstance(seller_info[0], dict) else ''
    #                        print(f"SELLER: {seller}")
    #            product_details = {
    #                'title': schema.get('name', ''),
    #                'images': [schema.get('image')] if schema.get('image') else [],
    #                'prices': [float(schema['offers']['price'].replace('$', ''))] if schema.get('offers', {}).get('price') else [],
    #                'currency': schema.get('offers', {}).get('priceCurrency', ''),
    #                'url': url,
    #                'description': schema.get('description', ''),
    #                'seller': seller.lower(),
    #            }
    #            parsed_products.append(product_details)
#
    #    return parsed_products

    
    
    def is_seller_verified(self, brand, seller):
        brand = brand.lower()
        seller = seller.lower()
        Logger.log('brand: {brand} seller: {seller}')
        return brand in seller or seller in brand
    
    
    
import threading   
class SearchEngine:
    def __init__(self, user_agents):
        self.user_agents = user_agents
        
    def create_brand_search_query(self, sku, brand_settings, iteration):
        if iteration <= 1:  # For the first two iterations, include brand and site operator
            domain = brand_settings.get("domain_hierarchy", [])[0] if brand_settings.get("domain_hierarchy") else ""
            query = f"site:{domain} \"{sku}\""
        else:  # For the rest, just use the SKU
            query = f"\"{sku}\""
        return f"https://www.google.com/search?q={query}"

    def choose_random_header(self):
        ua = random.choice(self.user_agents)
        #return ua.replace(";", "")
        return ua
 
    def filter_urls_by_brand_and_whitelist(self, urls, brand_settings, whitelisted_domains):
        brand_domains = [domain.replace('www.', '') for domain in brand_settings.get("domain_hierarchy", [])]
        whitelisted_domains = [domain.replace('www.', '') for domain in whitelisted_domains]
        approved_brand_urls = []
        approved_whitelist_urls = []

        if isinstance(urls, str):
            urls = urls.split(',')

        for url in urls:
            url = str(url).strip()
            if not url.startswith(('http://', 'https://')):
                url = 'http://' + url

            try:
                parsed_url = urlparse(url)
                domain = parsed_url.netloc
                if domain.startswith('www.'):
                    domain = domain[4:]

                if domain in brand_domains:
                    approved_brand_urls.append([url, "brand"])
                elif domain in whitelisted_domains:
                    approved_whitelist_urls.append([url, "whitelist"])

            except Exception as e:
                Logger.log(f"Error parsing URL '{url}': {e}")
        
        # Combine brand URLs and whitelisted URLs
        approved_urls = approved_brand_urls + approved_whitelist_urls
        return approved_urls if approved_urls else None
################################
    def filter_urls_by_currency(self, currency_items, urls):
        filtered_urls = []
        Logger.log(f'Filtered {urls}')
        for url in urls:
            Logger.log(f"url: {url}")
            for item in currency_items:
                #print(f'item: {item} url: {url}')
                #print(f'item: {type(item)} url: {type(url)}')
                #print(url)
                if str(item.lower()) in str(url[0]).lower():
                    Logger.log(f"item: {item} url: {url}")
                    filtered_urls.append(url)
                    break
            
        return filtered_urls
###############################

import threading
import requests
import random
import time

class Azure:
    active_azure_urls = []  # Shared among all instances
    lock = threading.Lock()  # Lock for thread-safe operations
    inactive_azure_urls = []  # Inactive URLs shared among all instances

    def __init__(self, azure_urls, user_agents):
        self.user_agents = user_agents
        self.request_count = 0
        self.index = 0
        with Azure.lock:
            if not Azure.active_azure_urls:  # Initialize only once
                Azure.active_azure_urls = azure_urls.copy()
        self.azure_checker_thread = threading.Thread(target=self.check_inactive_urls)
        self.azure_checker_thread.daemon = True
        self.azure_checker_thread.start()

    def choose_random_header(self):
        ua = random.choice(self.user_agents)
        return ua

    def fetch_target_body_azure(self, each_sku, query):
        user_agent = self.choose_random_header()
        self.request_count += 1

        while not Azure.active_azure_urls:  # wait until there is at least one active Azure URL
            Logger.log("No active Azure URLs. Waiting for one to become active...")
            time.sleep(60)  # wait for 60 seconds before checking again

        with Azure.lock:
            Logger.log(f"THIS IS THE REQUEST COUNT: {self.request_count}")
            if self.request_count > 100:
                self.index += 1  # Go to next azure url
                self.index = self.index % len(Azure.active_azure_urls)  # Loop back to first azure url if at end
                self.request_count = 0  # Reset request count
                if self.index==len(Azure.active_azure_urls)-1:
                    time.sleep(30)
                    Logger.log("Sleeping for 30 seconds")
            if self.index < len(Azure.active_azure_urls):
                azure_url = Azure.active_azure_urls[self.index]  # Get new azure url
            else:
                Logger.log("stuck in infinite loop")
                return self.fetch_target_body_azure(each_sku, query)
            Logger.log(f"Using Azure URL: {azure_url}")

        surl = f"{azure_url}&id={each_sku}&header={user_agent}&url={query}"
        try:
            response = requests.get(surl)
            response.raise_for_status()
            time.sleep(30)
            Logger.log(f"Status Code:{response.status_code}")
            return response.text if response else ""
        except requests.RequestException as e:
            Logger.log(f"An error occurred: {e}, for url {surl}, for sku {each_sku}, for query {query}")
            if "429" in str(e) or "500" in str(e):
                self.deactivate_url(azure_url)
                self.start_cooldown_timer(azure_url)
                return self.fetch_target_body_azure(each_sku, query)
            elif "404" in str(e) and "https://webcache.googleusercontent.com/search?q=cache:" not in query and "https://www.google.com/search?q=" not in query:
                cached_query = f"https://webcache.googleusercontent.com/search?q=cache:{query}"
                return self.fetch_target_body_azure(each_sku, cached_query)
            


    def deactivate_url(self, url):
        with Azure.lock:
            if url in Azure.active_azure_urls:
                Azure.active_azure_urls.remove(url)
                Azure.inactive_azure_urls.append(url)

    def start_cooldown_timer(self, url, cooldown_time=60):
        timer = threading.Timer(cooldown_time, self.reactivate_url, [url])
        timer.start()

    def reactivate_url(self, url):
        with Azure.lock:
            if url in Azure.inactive_azure_urls:
                Azure.inactive_azure_urls.remove(url)
                Azure.active_azure_urls.append(url)
                
    def check_inactive_urls(self, check_interval=300):  # Check every 5 minutes
        while True:
            with Azure.lock:
                for url in Azure.inactive_azure_urls:
                    self.test_inactive_url(url)
            time.sleep(check_interval)

    def test_inactive_url(self, url):
        try:
            response = requests.get(url)
            response.raise_for_status()
            with Azure.lock:
                if url in Azure.inactive_azure_urls:
                    Azure.inactive_azure_urls.remove(url)
                    Azure.active_azure_urls.append(url)
            Logger.log(f"URL {url} is back online and moved to active Azure URLs.")
        except requests.RequestException as e:
            threading.Timer(60, self.test_inactive_url, args=[url]).start()





class DataHandler:
    @staticmethod
    def read_input_data(file_path):
        input_data = []
        with open(file_path, mode='r', encoding='utf-8') as file:
            reader = csv.reader(file, delimiter='\t')
            for row in reader:
                if row:  # Check if row is not empty
                    input_data.append({'sku': row[0], 'brand': row[1]})
        return input_data

    @staticmethod
    def write_output_data(product, file_path):
        with open(file_path, mode='a', encoding='utf-8', newline='') as file:
            writer = csv.writer(file, delimiter='\t')

            # Correctly handle nested lists in 'images'
            if product.images and all(isinstance(item, list) for item in product.images):
                images_flat = ','.join([img for sublist in product.images for img in sublist])
            else:
                images_flat = ','.join(product.images) if product.images else ''
            row = [
                product.input_sku,
                product.source_type,
                product.currency,
                ','.join(map(str, product.prices)),
                product.url,
                product.title,
                images_flat,
                product.description,
                product.seller
            ]
            writer.writerow(row)
        
class ExcelProcessor():
    
    def __init__(self,filepath, searchcol,brandcol,destcol, preprocessing_option="append", min_row=6):
        self.filepath = filepath
        self.searchcol = searchcol
        self.brandcol = brandcol
        self.destcol = destcol
        self.preprocessing_option = preprocessing_option
        self.min_row = min_row 
        self.wb=load_workbook(self.filepath)
        self.ws = self.wb.active
        
        
        #Create new columns
        self.make_new_col("Source URL")
        self.make_new_col("Source Type")
        self.make_new_col("Seller")
    
    def read_excel(self):
        excelValues=[]
        # Iterate through the rows, considering only those without a corresponding value in destcol
        for index,row in enumerate(self.ws.iter_rows(min_row=self.min_row, values_only=True)):
                search_value=row[self.searchcol]
                dest_value = row[self.destcol]
                brand_value=row[self.brandcol]
                if (search_value is not None) and (dest_value is None or dest_value == '') and (self.preprocessing_option == "append"):
                    excelValues.append({'sku': search_value, 'brand': brand_value, 'excel_row_number':index+self.min_row+1})
                elif(self.preprocessing_option != "append"):
                    excelValues.append({'sku': search_value, 'brand': brand_value,'excel_row_number':index+self.min_row+1})    
        return excelValues

    #def write_excel(self, product):
    #    for row_number, row in enumerate(self.ws.iter_rows(min_row=self.min_row, values_only=True), start=self.min_row):
    #        if row[self.searchcol] == product.input_sku:  # Adjusted index to 0-based
    #            self.ws.cell(row=row_number, column=self.destcol+1).value = str(product.prices)
    #            print(product.prices)
    #            
    #            # Insert a new column for product.url
    #            self.ws.cell(row=row_number, column=self.find_col_index("Source URL")).value = str(product.url)
    #            
    #            # Insert another new column for product.source_type
    #            self.ws.cell(row=row_number, column=self.find_col_index("Source Type")).value = str(product.source_type)
    #            
    #            
    #            self.ws.cell(row=row_number, column=self.find_col_index("Seller")).value = str(product.seller)
    #            
    #            
    #            self.wb.save(self.filepath)
    #            return True
    #    return False
    def write_excel(self, output):
        for product_info in output:
            self.ws.cell(row=int(product_info[0])-1, column=self.destcol + 1).value = product_info[1]
            self.ws.cell(row=int(product_info[0])-1, column=self.find_col_index("Source URL")).value = product_info[2]
            self.ws.cell(row=int(product_info[0])-1, column=self.find_col_index("Source Type")).value = product_info[3]
            self.ws.cell(row=int(product_info[0])-1, column=self.find_col_index("Seller")).value = product_info[4]
        self.wb.save(self.filepath)
       
    
    
    def make_new_col(self, col_name):
        header_row = self.ws[self.min_row-1]
        if col_name not in [cell.value for cell in header_row]:
            max_col=self.ws.max_column
            self.ws.insert_cols(max_col+1)
            self.ws.cell(row=self.min_row-1, column=max_col+1).value=col_name
            self.wb.save(self.filepath)
            
    def find_col_index(self, col_name):
        for cell in self.ws.iter_rows(min_row=self.min_row-1, max_row=self.min_row-1, values_only=True):
            for index, value in enumerate(cell):
                if value == col_name:
                    return index + 1 # Adjusted index to 0-based
        return None
class ProductSchema:
    def __init__(self,product_schemas, source=False):
        self.product_schemas = product_schemas
        self.source=source
        self.parsed_products = self.parse_product_schemas(self.product_schemas)
        
        
    def get_parsed_products(self):
        return self.parsed_products

    def parse_product_schemas(self,product_schemas):
        parsed_products = []

        for schema in product_schemas:
            if schema.get('@type') == 'Product':
                offers_info = self.extract_offers(schema)
                for offer in offers_info:
                    
                    if(offer.get('@type') == 'Offer'):
                        prices=self.get_prices(offer)
                        currency=self.get_currency(offer)
                        seller=self.get_seller(offer)
                        description=self.get_description(offer)
                        title=self.get_title(offer)
                        images=self.get_images(offer)
                        url=self.get_url(offer)
                        product_details = self.create_product_details(title,images,prices,currency,url,description,seller,schema)
                        parsed_products.append(product_details)
                        
                    elif(offer.get('@type') == 'AggregateOffer'):
                        for suboffer in self.extract_offers(offer):
                            prices=self.get_prices(suboffer)
                            currency=self.get_currency(suboffer)
                            seller=self.get_seller(suboffer)
                            description=self.get_description(suboffer)
                            title=self.get_title(suboffer)
                            images=self.get_images(suboffer)
                            url=self.get_url(suboffer)
                            product_details = self.create_product_details(title,images,prices,currency,url,description,seller,schema)
                            parsed_products.append(product_details)
        return parsed_products



    def get_title(self, data):
        if isinstance(data, dict):
            for key, value in data.items():
                if key.lower() not in ['seller','brand']:
                    if key == 'name':
                        return value
                    else:
                        result = self.get_title(value)
                        if result:
                            return result
        else: return None        
            
    def get_images(self,data):
        images = []
        if isinstance(data, dict):
            for key, value in data.items():
                if key == 'image' and isinstance(value, (list, str)):
                    if isinstance(value, list):
                        images.extend(value)
                    else:
                        images.append(value)
                else:
                    images.extend(self.get_images(value))
        elif isinstance(data, list):
            for item in data:
                images.extend(self.get_images(item))
        return images

    def get_prices(self,data):
        prices = []
        if isinstance(data, dict):
            for key, value in data.items():
                if key.lower() in ['price', 'lowprice', 'highprice'] and isinstance(value, str):
                    prices.append(value.replace("$", "").replace(",", "").replace(" ", ""))
                elif key.lower() in ['price', 'lowprice', 'highprice'] and isinstance(value, (int, float)):
                    prices.append(value)
                else:
                    prices.extend(self.get_prices(value))
        elif isinstance(data, list):
            for item in data:
                prices.extend(self.get_prices(item))
        return prices

    def get_currency(self,data):
        if isinstance(data, dict):
            currency = data.get('priceCurrency', None)
            if currency:
                return currency
            for value in data.values():
                result = self.get_currency(value)
                if result:
                    return result
        elif isinstance(data, list):
            for item in data:
                result = self.get_currency(item)
                if result:
                    return result
    def get_url(self,data):
        if self.source:
            if isinstance(data, dict):
                url = data.get('url', None)
                if url:
                    return f"https://modesens.com{url}"
                for value in data.values():
                    result = self.get_url(value)
                    if result:
                        return f"https://modesens.com{url}"
            elif isinstance(data, list):
                for item in data:
                    result = self.get_url(item)
                    if result:
                        return f"https://modesens.com{url}"
        else:
            if isinstance(data, dict):
                url = data.get('url', None)
                if url:
                    return f"{url}"
                for value in data.values():
                    result = self.get_url(value)
                    if result:
                        return f"{url}"
            elif isinstance(data, list):
                for item in data:
                    result = self.get_url(item)
                    if result:
                        return f"{url}"
        
                
                
    def get_description(self,data):
        if isinstance(data, dict):
            for key, value in data.items():
                if key == 'description':
                    return value
                else:
                    result = self.get_description(value)
                    if result:
                        return result
                    
    def get_seller(self,data):
        if isinstance(data, dict):
            seller = data.get('seller', None)
            if isinstance(seller, dict) and 'name' in seller:
                return seller['name']
            for value in data.values():
                result = self.get_seller(value)
                if result:
                    return result
        elif isinstance(data, list):
            for item in data:
                result = self.get_seller(item)
                if result:
                    return result


    def extract_offers(self,data):
        offers = []

        if isinstance(data, dict):
            if 'offers' in data:
                # Directly append the offer or aggregate offer object
                offer_data = data['offers']
                if isinstance(offer_data, list):
                    offers.extend(offer_data)  # List of individual offers
                else:
                    offers.append(offer_data)  # Single or aggregate offer
            else:
                # Recursively search for offers in other dictionary values
                for value in data.values():
                    offers.extend(self.extract_offers(value))

        elif isinstance(data, list):
            # If the data is a list, apply the function to each element
            for item in data:
                offers.extend(self.extract_offers(item))

        return offers

    def create_product_details(self, title,images,prices,currency,url,description,seller,schema):
        product_details = {
                        'title': title,  
                        'images': images,  
                        'prices': prices,
                        'currency': currency,
                        'url': url,  
                        'description': description,  
                        'seller': seller.lower() if seller else None
                    }
        for key, value in product_details.items():
            if value in [None,[],"",{}]:
                if key == 'title':
                    product_details[key] = self.get_title(schema)
                elif key == 'images':
                    product_details[key] = self.get_images(schema)
                elif key == 'prices':
                    product_details[key] = self.get_prices(schema)
                elif key == 'currency':
                    product_details[key] = self.get_currency(schema)
                elif key == 'url':
                    product_details[key] = self.get_url(schema)
                elif key == 'description':
                    product_details[key] = self.get_description(schema)
                elif key == 'seller':
                    seller = self.get_seller(schema)
                    product_details[key] = seller.lower() if seller else seller
        return product_details
    
    
 
class Logger():
    def __init__(self, file_name):
        self.input_file_name=file_name
        self.logger=self.setup_logger()
    
    
    
    def setup_logger(self):
        logger = logging.getLogger('DataLogger')
        logger.setLevel(logging.INFO)  # Set the logging level

        # Create file handler which logs even debug messages
        timestamp = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
        log_file_path = f"{self.input_file_name}_{timestamp}_data_log.log"
        
        fh = logging.FileHandler(log_file_path)
        fh.setLevel(logging.INFO)

        # Create formatter and add it to the handlers
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
        fh.setFormatter(formatter)

        # Add the handlers to the logger
        logger.addHandler(fh)

        return logger
    
    @staticmethod
    def log_product(product):
        dict={}
        logger = logging.getLogger('DataLogger')
        dict['sku']=product.input_sku
        dict['source_type']=product.source_type
        dict['currency']=product.currency
        dict['prices']=product.prices
        dict['url']=product.url
        dict['title']=product.title
        dict['images']=product.images
        dict['description']=product.description
        dict['seller']=product.seller
        dict['excel_row_number']=product.excel_row_number
        logger.info(f"{product.input_sku}Product: {dict}")
        
    @staticmethod
    def log(message):
        logger = logging.getLogger('DataLogger')
        logger.info(message)
        
    
    
